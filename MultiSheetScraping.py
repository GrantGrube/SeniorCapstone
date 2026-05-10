from openpyxl import load_workbook
import json

book = load_workbook(filename="CapstoneSheet.xlsx")

workbook_data = {"sheets": []}

for sheet in book.worksheets:
    sheet_data = {
        "sheetName": sheet.title,
        "rows": []
    }

    for row in sheet.iter_rows(values_only=True):
        row_data = {"cells": []}

        for cell in row:
            row_data["cells"].append({
                "value": str(cell) if cell is not None else "",
                "type": type(cell).__name__ if cell is not None else "empty"
            })

        sheet_data["rows"].append(row_data)

    workbook_data["sheets"].append(sheet_data)

output_path = r"C:\Users\Grant\Desktop\Capstone\UnityCapstone\Assets\StreamingAssets\excel_data.json"
with open(output_path, "w") as f:
    json.dump(workbook_data, f, indent=2)

print("JSON formatting complete, please check Asset file in Unity")