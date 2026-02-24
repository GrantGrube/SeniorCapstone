import openpyxl
import json
from openpyxl import workbook
from openpyxl import load_workbook

book = load_workbook(filename="Capstone Sheet.xlsx") #loading the sheet we want to use

print(book.sheetnames)
sheetn = input("Please enter the name of the sheet you would like to use: ")
current = book[sheetn] #change current sheet to the one selected
print("you have chosen " + current.title)


cell_data = [] #initialize the array for each cell


for row in current.iter_rows():
    for cell in row:
        cell_data.append({     #adding specific data from each cell
            "row": cell.row,
            "col": cell.column,
            "coordinate": cell.coordinate,
            "value": cell.value,
            "type": type(cell.value).__name__
        })

#exporting directly to Unity Asset file 
output_path = r"\Users\Grant\Desktop\Capstone\UnityCapstone\Assets\StreamingAssets/excel_data.json"
with open(output_path, "w") as f:
    json.dump(cell_data, f, indent=4)

print("JSON formatting complete, please check Asset file in Unity")

